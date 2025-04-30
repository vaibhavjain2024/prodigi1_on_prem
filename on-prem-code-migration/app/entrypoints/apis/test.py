from fastapi.routing import APIRouter
from fastapi.responses import StreamingResponse
import io
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from utils.logger import get_logger
logger = get_logger("testlog", "testendpointlogs")


router = APIRouter(prefix="/test")

# Sample data for multiple sheets
data1 = [
    {"Name": "Alice", "Age": 30, "City": "New York"},
    {"Name": "Bob", "Age": 25, "City": "Los Angeles"},
    {"Name": "Charlie", "Age": 35, "City": "Chicago"},
]

data2 = [
    {"Product": "Laptop", "Price": 1000, "Stock": 50},
    {"Product": "Phone", "Price": 500, "Stock": 200},
    {"Product": "Tablet", "Price": 750, "Stock": 150},
]

def generate_excel(data1, data2):
    logger.info("Generating Excel")
    logger.info("Excel Generated")
    # Create a workbook
    wb = Workbook()

    # Add first sheet (data1)
    sheet1 = wb.active
    sheet1.title = "Sheet1"  # You can name this sheet whatever you like
    headers1 = list(data1[0].keys())
    sheet1.append(headers1)

    for row in data1:
        sheet1.append(list(row.values()))

    # Adjust column widths for the first sheet
    for col_num, col_name in enumerate(headers1, start=1):
        max_length = 0
        column = get_column_letter(col_num)
        for row in sheet1.iter_rows(min_col=col_num, max_col=col_num):
            for cell in row:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
        adjusted_width = (max_length + 2)
        sheet1.column_dimensions[column].width = adjusted_width

    # Add second sheet (data2)
    sheet2 = wb.create_sheet("Sheet2")  # Create a new sheet
    headers2 = list(data2[0].keys())
    sheet2.append(headers2)

    for row in data2:
        sheet2.append(list(row.values()))

    # Adjust column widths for the second sheet
    for col_num, col_name in enumerate(headers2, start=1):
        max_length = 0
        column = get_column_letter(col_num)
        for row in sheet2.iter_rows(min_col=col_num, max_col=col_num):
            for cell in row:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
        adjusted_width = (max_length + 2)
        sheet2.column_dimensions[column].width = adjusted_width

    # Save the workbook to a BytesIO object
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    return file_stream

@router.get("/download_excel/")
async def download_excel():
    logger.info("Calling Download excel")
    # Generate Excel file from the list of dictionaries
    excel_file = generate_excel(data1, data2)

    # Return the file as a StreamingResponse
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=data.xlsx"}
    )
