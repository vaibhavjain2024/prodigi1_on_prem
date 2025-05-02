from fastapi import Depends, HTTPException
from fastapi.routing import APIRouter
from fastapi.responses import StreamingResponse, JSONResponse

from io import StringIO, BytesIO
from csv import DictWriter
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from app.schema.reportsSchema import (
    Reports, reportFilter
)

from app.onPremServices.reports import (
    msil_iot_psm_get_report_filters, 
    msil_iot_psm_get_reports_downtime, msil_iot_psm_get_reports_quality, msil_iot_psm_get_reports_production,
    msil_iot_psm_get_reports_downtime_report, msil_iot_psm_get_reports_quality_report, msil_iot_psm_get_reports_production_report
)

def generate_excel(data_dict):
    # Create a workbook
    wb = Workbook()
    
    # Remove the default sheet created
    wb.remove(wb.active)

    # Iterate over each dataset in the dictionary and create a sheet for each
    for sheet_name, data_list in data_dict.items():
        sheet = wb.create_sheet(sheet_name)  # Create a new sheet for each key
        headers = list(data_list[0].keys())  # Use the keys of the first dictionary as headers
        sheet.append(headers)

        # Add the rows to the sheet
        for row in data_list:
            sheet.append(list(row.values()))

        # Adjust column widths for the current sheet
        for col_num, col_name in enumerate(headers, start=1):
            max_length = 0
            column = get_column_letter(col_num)
            for row in sheet.iter_rows(min_col=col_num, max_col=col_num):
                for cell in row:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column].width = adjusted_width

    # Save the workbook to a BytesIO object
    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    return file_stream


router = APIRouter(prefix="/pressShop/reports")

def returnJsonResponse(response):
    return JSONResponse(content=response, status_code=200)

@router.get("/filters")
def report_Filter(reportfilter: reportFilter = Depends()):
    shop_id = reportfilter.shop_id
    if not shop_id:
        raise HTTPException(status_code=400, detail="Missing 'shop_id' query parameter")
    
    response = msil_iot_psm_get_report_filters.handler(shop_id)
    return returnJsonResponse(response)

@router.get("/downtime")
def get_Downtime(reports: Reports = Depends()):
    shop_id = reports.shop_id
    start_date = reports.start_date
    end_date = reports.end_date
    if not shop_id or not start_date or not end_date:
        raise HTTPException(status_code=400, detail="Missing 'shop_id' / 'start_date' / 'end_date' query parameter")

    query_params = reports.model_dump(exclude={"shop_id", "start_date", "end_date"}, exclude_none=True)
    response = msil_iot_psm_get_reports_downtime.handler(shop_id, start_date, end_date, **query_params)
    return returnJsonResponse(response)

@router.get("/quality")
def get_Quality(reports: Reports = Depends()):
    shop_id = reports.shop_id
    start_date = reports.start_date
    end_date = reports.end_date
    if not shop_id or not start_date or not end_date:
        raise HTTPException(status_code=400, detail="Missing 'shop_id' / 'start_date' / 'end_date' query parameter")

    query_params = reports.model_dump(exclude={"shop_id", "start_date", "end_date"}, exclude_none=True)
    response = msil_iot_psm_get_reports_quality.handler(shop_id, start_date, end_date, **query_params)
    return returnJsonResponse(response)

@router.get("/production")
def get_Production(reports: Reports = Depends()):
    shop_id = reports.shop_id
    start_date = reports.start_date
    end_date = reports.end_date
    if not shop_id or not start_date or not end_date:
        raise HTTPException(status_code=400, detail="Missing 'shop_id' / 'start_date' / 'end_date' query parameter")

    query_params = reports.model_dump(exclude={"shop_id", "start_date", "end_date"}, exclude_none=True)
    response = msil_iot_psm_get_reports_production.handler(shop_id, start_date, end_date, **query_params)
    return returnJsonResponse(response)



@router.get("/downtime/report")
def get_Downtime_Reports(reports: Reports = Depends()):
    shop_id = reports.shop_id
    start_date = reports.start_date
    end_date = reports.end_date
    if not shop_id or not start_date or not end_date:
        raise HTTPException(status_code=400, detail="Missing 'shop_id' / 'start_date' / 'end_date' query parameter")

    query_params = reports.model_dump(exclude={"shop_id", "start_date", "end_date"}, exclude_none=True)
    response = msil_iot_psm_get_reports_downtime_report.handler(shop_id, start_date, end_date, **query_params)

    result = StringIO()
    fieldnames = response[0].keys()

    writer = DictWriter(result, fieldnames=fieldnames)
    writer.writeheader()
    writer.writerows(response)

    result.seek(0)

    return StreamingResponse(result, media_type="text/csv", headers={"Content-Disposition": "attachment; filename=downtime_report.csv"})

@router.get("/quality/report")
def get_Quality_Reports(reports: Reports = Depends()):
    shop_id = reports.shop_id
    start_date = reports.start_date
    end_date = reports.end_date
    if not shop_id or not start_date or not end_date:
        raise HTTPException(status_code=400, detail="Missing 'shop_id' / 'start_date' / 'end_date' query parameter")

    query_params = reports.model_dump(exclude={"shop_id", "start_date", "end_date"}, exclude_none=True)
    response = msil_iot_psm_get_reports_quality_report.handler(shop_id, start_date, end_date, **query_params)

    result = StringIO()
    fieldnames = response[0].keys()

    writer = DictWriter(result, fieldnames=fieldnames)
    writer.writeheader()
    writer.writerows(response)

    result.seek(0)

    return StreamingResponse(result, media_type="text/csv", headers={"Content-Disposition": "attachment; filename=quality_report.csv"})

@router.get("/production/report")
def get_Production_Reports(reports: Reports = Depends()):
    shop_id = reports.shop_id
    start_date = reports.start_date
    end_date = reports.end_date
    if not shop_id or not start_date or not end_date:
        raise HTTPException(status_code=400, detail="Missing 'shop_id' / 'start_date' / 'end_date' query parameter")

    query_params = reports.model_dump(exclude={"shop_id", "start_date", "end_date"}, exclude_none=True)
    response = msil_iot_psm_get_reports_production_report.handler(shop_id, start_date, end_date, **query_params)

    excel_file_data = generate_excel(response)
    return StreamingResponse(
        excel_file_data, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=production_report.xlsx"}
    )