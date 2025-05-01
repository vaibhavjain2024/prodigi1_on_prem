import io
import pandas as pd
from openpyxl import Workbook, load_workbook

from . import aws_utils


def get_file_body_and_sheet_names_of_xlsx_file(bucket, master_file_path):
    file_body = aws_utils.read_file_from_s3_bucket(bucket, master_file_path)
    workbook = load_workbook(io.BytesIO(file_body), data_only=True)
    get_sheet_names = workbook.sheetnames
    return file_body, get_sheet_names


def read_xlsx_file_sheet_and_remove_description(
    file_body, sheet_name, header_row=[0], col_indexes=0
):
    recv_sheet_data = pd.read_excel(file_body, sheet_name=sheet_name, header=header_row)
    sheet_data = recv_sheet_data.drop(
        [recv_sheet_data.index[i] for i in range(col_indexes)]
    )
    return sheet_data
