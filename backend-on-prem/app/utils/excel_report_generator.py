# utils/excel_report.py

from openpyxl import Workbook
from openpyxl.styles import Font
from io import BytesIO


def excel_report_generator(
    transformed_data: list,
    excel_headers: list,
    metadata_kv_rows: list,
    report_title: str,
    sheet_title: str = "Report"
) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title
    bold = Font(bold=True)
    if report_title:
        ws.append(["Report title", report_title])
        ws.cell(row=ws.max_row, column=1).font = bold
        ws.append([])

    # Add metadata rows
    if metadata_kv_rows:
        for item in metadata_kv_rows:
            label = item.get("label", "")
            value = item.get("value", "")
            ws.append([f"{label}:", value])
            if label:
                ws.cell(row=ws.max_row, column=1).font = bold

    # Add spacing
    ws.append([])
    ws.append([])

    # Header row
    ws.append(excel_headers)
    for cell in ws[ws.max_row]:
        cell.font = bold

    # Data rows
    for row in transformed_data:
        ws.append([row.get(h, "") for h in excel_headers])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
